#!/usr/bin/env pipenv-shebang
import os
import sqlite3
import datetime
import logging
import logging.config

import pretty_errors
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, numbers

import main as _
from sqlitehandler import create_connection, fetch_all


def create_report(today):
    params = (today + "%",)
    sql = """SELECT * FROM printrequest
        WHERE waktu LIKE ?
        ORDER BY petugas ASC, waktu ASC"""
    sql2 = """SELECT petugas, COUNT(petugas) FROM printrequest
        WHERE waktu LIKE ?
        GROUP BY petugas
        ORDER BY petugas ASC, waktu ASC"""

    conn = create_connection()
    if conn is not None:
        today2 = today.replace("-", "")
        filename = f"{today2}_report.xlsx"
        filepath = os.path.join(_.REPORT_DIR, filename)

        contents = fetch_all(conn, sql, params)
        contents2 = fetch_all(conn, sql2, params)

        if contents is not None and contents2 is not None:
            # create caption message
            i = 1
            j = 0
            row = []
            for name, count in contents2:
                name = " ".join(word.title() for word in name.split(" "))
                row.append(f"{i}. {name} ({count})")
                j += count
                i += 1
            caption = "\n".join(row)
            caption = f"{caption}\nTotal ({j})"

            # create xlsx file
            # ================================
            # style template header row
            header = NamedStyle(name="header")
            header.font = Font(name="Calibri", size=10, bold=True)
            header.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            # create a blank xlsx file
            wb = Workbook()
            ws = wb.active
            ws.merge_cells("A3:H3")

            # set column width
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 12.5
            ws.column_dimensions["F"].width = 20
            ws.column_dimensions["G"].width = 20
            ws.column_dimensions["H"].width = 25

            # set header title
            ws["A1"] = "PT POS INDONESIA (PERSERO)"
            ws["A1"].font = Font(name="Calibri", size=10)
            ws["A2"] = _.OFFICE
            ws["A2"].font = Font(name="Calibri", size=10)
            ws["A3"] = "LAPORAN CETAK PKB"
            ws["A3"].font = Font(bold=True)
            ws["A3"].alignment = Alignment(horizontal="center")
            ws["A5"] = "#"
            ws["B5"] = "NOPOL"
            ws["C5"] = "IDTRANS"
            ws["D5"] = "IDSAH"
            ws["E5"] = "BSU"
            ws["F5"] = "URL"
            ws["G5"] = "TANGGAL"
            ws["H5"] = "PETUGAS"

            # set style in header column
            header_row = ws[5]
            for cell in header_row:
                cell.style = header

            # style template body row
            uid = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
            body = NamedStyle(name=uid)
            body.font = Font(name="Calibri", size="10")

            # append data
            for content in contents:
                content = list(content)
                # add increment number
                i = ws[f"A{ws.max_row}"].value
                i = 1 if i == "#" else i + 1

                new_content = content.copy()
                new_content.insert(0, i)
                ws.append(new_content)

                # set style to body column
                body_row = ws[ws.max_row]
                for cell in body_row:
                    cell.style = body

            # set number format
            for cells in ws.iter_rows(min_col=5, max_col=5, min_row=6):
                for cell in cells:
                    cell.number_format = numbers.BUILTIN_FORMATS[43]

            # save to file
            wb.save(filepath)
            logger.info("file %s is created", filename)

            # close connection database
            conn.close()
        else:
            # contents is empty
            filename, filepath, caption = None, None, None
    else:
        # if no connection database
        filename, filepath, caption = None, None, None

    return filename, filepath, caption


if __name__ != "__main__":
    logger = logging.getLogger(__name__)
